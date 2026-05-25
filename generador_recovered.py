import re
import openpyxl
import glob
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import pdfplumber
except ImportError:
    print('Se requiere la librería pdfplumber. Instalando...')
    import subprocess
    import sys
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'pdfplumber'])
    import pdfplumber

def extract_account_number(pdf_path):
    '''Extrae el número de cuenta del PDF'''
    account_number = None
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages[:2]:
                text = page.extract_text()
                if not text:
                    continue
                match = re.search(r'\b(2341052324|2341055145)\b', text)
                if match:
                    account_number = match.group(1)
                    break
        return account_number
    except Exception as e:
        print(f'Error al extraer número de cuenta: {e}')
        return account_number

def extract_transactions_from_pdf(pdf_path):
    '''Extrae transacciones del PDF analizando el texto'''
    transactions = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f'PDF abierto. Total de páginas: {len(pdf.pages)}')
            last_saldo = None
            for page_num, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue
                lines = text.split('\n')
                saldo_anterior = None
                for line in lines:
                    line = line.lstrip('_').strip()
                    if 'SALDO ANTERIOR' in line:
                        saldo_match = re.search(r'([\d.]+,[0-9]{2})', line)
                        if saldo_match:
                            saldo_anterior = float(saldo_match.group(1).replace('.', '').replace(',', '.'))
                            last_saldo = saldo_anterior
                    
                    if re.match(r'\d{2}/\d{2}/\d{2}', line):
                        fecha = line[:8]
                        resto = line[8:].strip()
                        numeros_match = list(re.finditer(r'[\d.]+,[0-9]{2}', line))
                        if len(numeros_match) < 1:
                            continue
                        saldo_str = numeros_match[-1].group()
                        saldo = float(saldo_str.replace('.', '').replace(',', '.'))
                        
                        comprob_match = re.search(r'\s(\d+[A-Z0-9]*)\s+(?:[\d.]+,[0-9]{2})', resto)
                        if comprob_match:
                            comprob = comprob_match.group(1)
                            detalle_end = comprob_match.start()
                            detalle = resto[:detalle_end].strip()
                            resto_numeros = resto[comprob_match.start():]
                            montos = re.findall(r'[\d.]+,[0-9]{2}', resto_numeros)
                            debitos = 0
                            creditos = 0
                            if len(montos) >= 3:
                                val1 = float(montos[0].replace('.', '').replace(',', '.'))
                                val2 = float(montos[1].replace('.', '').replace(',', '.'))
                                if val1 > 0 and val2 == 0:
                                    debitos = val1
                                elif val2 > 0 and val1 == 0:
                                    creditos = val2
                                else:
                                    debitos = val1
                                    creditos = val2
                            elif len(montos) >= 2:
                                monto = float(montos[0].replace('.', '').replace(',', '.'))
                                if last_saldo is not None:
                                    if saldo > last_saldo:
                                        creditos = monto
                                    else:
                                        debitos = monto
                                else:
                                    debitos = monto
                            
                            if detalle and (debitos > 0 or creditos > 0):
                                transaction = {
                                    'FECHA': fecha,
                                    'DETALLE': detalle,
                                    'COMPROB.': comprob,
                                    'DEBITOS': debitos,
                                    'CREDITOS': creditos,
                                    'SALDO': saldo
                                }
                                transactions.append(transaction)
                            last_saldo = saldo
            print(f'\nTotal de transacciones extraídas: {len(transactions)}')
            if transactions:
                print(f'Primera transacción: {transactions[0]}')
            return transactions if transactions else None
    except FileNotFoundError:
        print(f'No se encontró el archivo PDF: {pdf_path}')
        return None
    except Exception as e:
        print(f'Error al procesar PDF: {e}')
        import traceback
        traceback.print_exc()
        return None

def create_excel_table_final(pdf_path):
    '''Crea un archivo Excel con formato de tabla final para el PDF dado'''
    print("Extrayendo transacciones...")
    transactions = extract_transactions_from_pdf(pdf_path)
    if not transactions:
        print("No se encontraron transacciones.")
        return
        
    account = extract_account_number(pdf_path) or "Cuenta_Desconocida"
    print(f"\nNúmero de cuenta detectado: {account}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"
    
    headers = ['FECHA', 'DETALLE', 'COMPROB.', 'DEBITOS', 'CREDITOS', 'SALDO']
    ws.append(headers)
    
    for t in transactions:
        ws.append([t['FECHA'], t['DETALLE'], t['COMPROB.'], t['DEBITOS'], t['CREDITOS'], t['SALDO']])
        
    # Formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Table logic
    tab = Table(displayName="EstadoCuenta", ref=f"A1:F{len(transactions)+1}")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # ---------------------------------------------------------
    # HOJA 2: Totales por Detalle
    # ---------------------------------------------------------
    ws_totales = wb.create_sheet("Totales por Detalle")
    totales_por_detalle = {}
    
    for t in transactions:
        detalle_original = t['DETALLE']
        detalle_lower = detalle_original.lower() if detalle_original else ""
        
        if 'debin' in detalle_lower:
            detalle = 'DEBIN'
        else:
            detalle = detalle_original
            
        debito = float(t['DEBITOS'])
        credito = float(t['CREDITOS'])
        
        if detalle not in totales_por_detalle:
            totales_por_detalle[detalle] = {'debitos': 0, 'creditos': 0, 'cantidad': 0}
            
        totales_por_detalle[detalle]['debitos'] += debito
        totales_por_detalle[detalle]['creditos'] += credito
        totales_por_detalle[detalle]['cantidad'] += 1

    totales_ordenados = sorted(totales_por_detalle.items(), key=lambda x: x[1]['cantidad'], reverse=True)
    
    headers_totales = ['DETALLE', 'CANTIDAD', 'TOTAL DEBITOS', 'TOTAL CREDITOS', 'SALDO NETO']
    ws_totales.append(headers_totales)
    
    for col_num, cell in enumerate(ws_totales[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, (det, tot) in enumerate(totales_ordenados, 2):
        saldo_neto = tot['creditos'] - tot['debitos']
        ws_totales.append([det, tot['cantidad'], tot['debitos'], tot['creditos'], saldo_neto])
        
    tab_tot = Table(displayName="TotalesDetalle", ref=f"A1:E{len(totales_ordenados)+1}")
    tab_tot.tableStyleInfo = style
    ws_totales.add_table(tab_tot)

    # ---------------------------------------------------------
    # HOJA 3: Resumen por Categoria
    # ---------------------------------------------------------
    ws_resumen = wb.create_sheet("Resumen por Categoria")
    
    # Categorías EXACTAS solicitadas y recuperadas del código original
    reglas_categorias = {
        'IMPUESTOS DEB/CRED': ['gravamenley25413sdeb', 'gravamenley25413scred', 'reintegroley25413deb', 'reintegroley25413cred', 'vsgytley25413'],
        'IVA DEBITO': ['ivabase'],
        'COMISIONES Y GASTOS': ['comispservrecaudacion', 'comispserv', 'comistransfne24', 'comiscompensacionatenc', 'comiscanjeobancos'],
        'CHEQUES DEBITADOS': ['pagochequepropiacasa'],
        'INTERDEPOSITOS': [
            'debtraninterblink', 'debtraninterblinkres', 'pagovepafip', 'pagoedensa', 
            'dbcredintranslinkcia', 'debitopagodirecto', 'pagomovistar', 'pagomovistarhogar', 
            'pagocajademedaportp', 'pagofaecys'
        ],
        'DEPOSITOS': [
            'debtraninterblinktit', 'rendpservrecaudacion', 'transfintdistlinklar', 
            'cbetrobcodsuc0001k', 'rendicionpagoslink', 'debin', 'transfintdistbanelar', 
            'crtrvariosvsuc0001', 'crtransfinterlinkres', 'dep', 'deposito', 'efectivo', 'ch',
            'transfintermmlinkular', 'camfeddistpzaobcos', '48hsbancos', 'crtransfinterbaneres',
            'creditosvarios', 'cobradoporcaja', 'crtrfacturasuc0001'
        ]
    }

    # Mantener el orden exacto solicitado
    orden_categorias = [
        'IMPUESTOS DEB/CRED',
        'IVA DEBITO',
        'COMISIONES Y GASTOS',
        'CHEQUES DEBITADOS',
        'INTERDEPOSITOS',
        'DEPOSITOS'
    ]

    resumen_totales = {cat: {'debitos': 0, 'creditos': 0, 'cantidad': 0} for cat in orden_categorias}

    for t in transactions:
        detalle = t['DETALLE'] or ""
        detalle_norm = re.sub(r'[^a-z0-9]', '', detalle.lower())
        debito = float(t['DEBITOS'])
        credito = float(t['CREDITOS'])
        
        for categoria, palabras_clave in reglas_categorias.items():
            if any(palabra in detalle_norm for palabra in palabras_clave):
                resumen_totales[categoria]['debitos'] += debito
                resumen_totales[categoria]['creditos'] += credito
                resumen_totales[categoria]['cantidad'] += 1
                break

    headers_resumen = ['CATEGORIA', 'CANTIDAD', 'TOTAL DEBITOS', 'TOTAL CREDITOS', 'SALDO NETO']
    ws_resumen.append(headers_resumen)
    
    for col_num, cell in enumerate(ws_resumen[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for cat in orden_categorias:
        tot = resumen_totales[cat]
        saldo_neto = tot['creditos'] - tot['debitos']
        ws_resumen.append([cat, tot['cantidad'], tot['debitos'], tot['creditos'], saldo_neto])
        row_idx += 1
            
    tab_res = Table(displayName="ResumenCategoria", ref=f"A1:E{row_idx-1}")
    tab_res.tableStyleInfo = style
    ws_resumen.add_table(tab_res)

    # Save
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    out_filename = f"estado_cuenta_final_{account}_{base_name}.xlsx"
    wb.save(out_filename)
    print(f"Archivo Excel con tabla final creado: {out_filename}")

if __name__ == '__main__':
    pdfs = glob.glob('Estado de Cuenta*.pdf')
    if not pdfs:
        pdfs = glob.glob('*.pdf')
    if not pdfs:
        print('No se encontraron archivos PDF para procesar en el directorio actual.')
    else:
        for pdf_file in pdfs:
            print(f'\n=== Procesando: {pdf_file} ===')
            create_excel_table_final(pdf_file)
